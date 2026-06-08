"""
Per-league results spreadsheet (.xlsx) builder.

The layout mirrors the shared league template: row 1 holds the league title plus
every member's name, row 2 holds each member's running total, and from row 3 on
there is one row per match. Columns A–D carry the fixture (home, away, actual
home, actual away); each member then owns three columns — predicted home,
predicted away, points.

The one rule worth stating loudly: a match's row stays blank (no result, picks
or points) until its lock time has passed, so a downloaded file can never leak an
upcoming pick. Reveal keys strictly off the lock time — not off whether a result
has been entered — so even an early-entered result can't expose a pick before the
match locks.

The sheet is colour-coded to match the shared league template (maroon title,
gold member headers, amber teams, navy results, grey points) with a frozen header
and an RTL view, so it's easy to read at a glance. The palette lives in consts.
"""
from decimal import Decimal
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import consts


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _member_columns(index):
    """Return (home_col, away_col, points_col) for the member at a 0-based index."""
    base = consts.EXPORT_FIRST_MEMBER_COL + index * consts.EXPORT_COLS_PER_MEMBER
    return base, base + 1, base + 2


def _team_label(team, label):
    """Persian display name for a fixture side (bracket slot if team is undecided)."""
    if team:
        return team.name_fa
    return consts.bracket_label_fa(label)


def _finish_layout(ws, border, member_count, match_count):
    """Apply the grid, column widths, frozen header, and RTL view to the sheet."""
    last_row = max(consts.EXPORT_TOTAL_ROW,
                   consts.EXPORT_FIRST_MATCH_ROW + match_count - 1)
    last_col = max(consts.EXPORT_COL_ACTUAL_AWAY,
                   consts.EXPORT_FIRST_MEMBER_COL
                   + member_count * consts.EXPORT_COLS_PER_MEMBER - 1)

    # Thin grid across the whole used range (borders empty cells too).
    for r in range(1, last_row + 1):
        for c in range(1, last_col + 1):
            ws.cell(row=r, column=c).border = border

    # Column widths: wide team names, narrow score cells.
    ws.column_dimensions[get_column_letter(consts.EXPORT_COL_HOME)].width = consts.EXPORT_WIDTH_TEAM
    ws.column_dimensions[get_column_letter(consts.EXPORT_COL_AWAY)].width = consts.EXPORT_WIDTH_TEAM
    ws.column_dimensions[get_column_letter(consts.EXPORT_COL_ACTUAL_HOME)].width = consts.EXPORT_WIDTH_RESULT
    ws.column_dimensions[get_column_letter(consts.EXPORT_COL_ACTUAL_AWAY)].width = consts.EXPORT_WIDTH_RESULT
    for i in range(member_count):
        h, a, p = _member_columns(i)
        ws.column_dimensions[get_column_letter(h)].width = consts.EXPORT_WIDTH_PRED
        ws.column_dimensions[get_column_letter(a)].width = consts.EXPORT_WIDTH_PRED
        ws.column_dimensions[get_column_letter(p)].width = consts.EXPORT_WIDTH_POINTS

    ws.row_dimensions[consts.EXPORT_TITLE_ROW].height = consts.EXPORT_TITLE_ROW_HEIGHT
    # Keep the title/total rows and the fixture columns on screen while scrolling.
    ws.freeze_panes = ws.cell(row=consts.EXPORT_FIRST_MATCH_ROW,
                              column=consts.EXPORT_FIRST_MEMBER_COL)
    ws.sheet_view.rightToLeft = consts.EXPORT_STYLE_RTL


def build_league_workbook(league, now=None):
    """Build and return an openpyxl Workbook of the league's results."""
    from .models import Match, MatchScore, Membership, Prediction

    now = now or timezone.now()

    memberships = list(
        Membership.objects.filter(league=league)
        .select_related("user")
        .order_by("joined_at")  # stable column order: members never reshuffle
    )
    member_index = {m.id: i for i, m in enumerate(memberships)}

    matches = list(
        Match.objects.filter(competition=league.competition)
        .select_related("home_team", "away_team")
        .order_by("kickoff", "match_number")
    )

    # A match is "revealed" strictly once its lock time has passed. Keying off the
    # lock time alone (not match.is_finished) means that even if a result is entered
    # early — before the match would normally lock — its picks, result and points
    # all stay hidden until lock. That's the rule that makes an upcoming prediction
    # truly invisible in the downloaded file.
    revealed_ids = {
        m.id for m in matches if now >= m.lock_time(league.lock_minutes)
    }

    # predictions[match_id][membership_id] -> Prediction
    predictions = {}
    for p in Prediction.objects.filter(membership__league=league):
        predictions.setdefault(p.match_id, {})[p.membership_id] = p

    # scores[match_id][membership_id] -> MatchScore, and each member's total.
    # Totals only sum scores from *revealed* matches, so the printed total always
    # equals the sum of the points cells actually shown below it (no hidden leak).
    scores = {}
    totals = {m.id: Decimal("0") for m in memberships}
    for s in MatchScore.objects.filter(membership__league=league):
        scores.setdefault(s.match_id, {})[s.membership_id] = s
        if s.match_id in revealed_ids:
            totals[s.membership_id] = totals.get(s.membership_id, Decimal("0")) + s.points

    wb = Workbook()
    ws = wb.active
    ws.title = (league.name or "")[:consts.EXPORT_SHEET_TITLE_MAX] or consts.EXPORT_SHEET_TITLE_FALLBACK

    # --- Styles (built from consts so the palette lives in one place) --------- #
    title_fill = _fill(consts.EXPORT_COLOR_TITLE_BG)
    title_font = Font(bold=True, size=consts.EXPORT_TITLE_FONT_SIZE, color=consts.EXPORT_COLOR_TITLE_FG)
    header_fill = _fill(consts.EXPORT_COLOR_HEADER_BG)
    header_font = Font(bold=True, size=consts.EXPORT_HEADER_FONT_SIZE, color=consts.EXPORT_COLOR_HEADER_FG)
    total_fill = _fill(consts.EXPORT_COLOR_TOTAL_BG)
    total_font = Font(bold=True, size=consts.EXPORT_TOTAL_FONT_SIZE, color=consts.EXPORT_COLOR_TOTAL_FG)
    team_fill = _fill(consts.EXPORT_COLOR_TEAM_BG)
    team_font = Font(size=consts.EXPORT_BODY_FONT_SIZE, color=consts.EXPORT_COLOR_TEAM_FG)
    result_fill = _fill(consts.EXPORT_COLOR_RESULT_BG)
    result_font = Font(bold=True, size=consts.EXPORT_BODY_FONT_SIZE, color=consts.EXPORT_COLOR_RESULT_FG)
    pred_fill = _fill(consts.EXPORT_COLOR_PRED_BG)
    pred_font = Font(size=consts.EXPORT_BODY_FONT_SIZE, color=consts.EXPORT_COLOR_PRED_FG)
    points_fill = _fill(consts.EXPORT_COLOR_POINTS_BG)
    points_font = Font(bold=True, size=consts.EXPORT_POINTS_FONT_SIZE, color=consts.EXPORT_COLOR_POINTS_FG)
    side = Side(style="thin", color=consts.EXPORT_COLOR_BORDER)
    border = Border(left=side, right=side, top=side, bottom=side)
    center = Alignment(horizontal="center", vertical="center")
    start = Alignment(vertical="center")  # "general": text reads from the RTL edge

    def put(r, c, value, fill=None, font=None, align=center):
        cell = ws.cell(row=r, column=c, value=value)
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font
        cell.alignment = align
        return cell

    fixture_cols = range(consts.EXPORT_COL_AWAY, consts.EXPORT_FIRST_MEMBER_COL)  # B,C,D

    # Row 1: maroon title banner across the fixture columns + gold member tabs.
    put(consts.EXPORT_TITLE_ROW, consts.EXPORT_COL_HOME, league.name, title_fill, title_font, start)
    for c in fixture_cols:
        put(consts.EXPORT_TITLE_ROW, c, None, title_fill)
    # Row 2: pale-gold standings band — a label over the fixture, totals per member.
    put(consts.EXPORT_TOTAL_ROW, consts.EXPORT_COL_HOME, consts.EXPORT_LABEL_TOTAL, total_fill, total_font, start)
    for c in fixture_cols:
        put(consts.EXPORT_TOTAL_ROW, c, None, total_fill)

    for m in memberships:
        h, a, p = _member_columns(member_index[m.id])
        put(consts.EXPORT_TITLE_ROW, h, m.user.public_name, header_fill, header_font)
        put(consts.EXPORT_TITLE_ROW, a, None, header_fill)
        put(consts.EXPORT_TITLE_ROW, p, None, header_fill)
        put(consts.EXPORT_TOTAL_ROW, h, float(totals[m.id]), total_fill, total_font)
        put(consts.EXPORT_TOTAL_ROW, a, None, total_fill)
        put(consts.EXPORT_TOTAL_ROW, p, None, total_fill)

    # One row per match, in kickoff order.
    row = consts.EXPORT_FIRST_MATCH_ROW
    for match in matches:
        put(row, consts.EXPORT_COL_HOME,
            _team_label(match.home_team, match.home_label), team_fill, team_font, start)
        put(row, consts.EXPORT_COL_AWAY,
            _team_label(match.away_team, match.away_label), team_fill, team_font, start)

        # Not locked yet -> show only the fixture; no result, picks or points leak.
        if match.id not in revealed_ids:
            row += 1
            continue

        if match.is_finished:
            put(row, consts.EXPORT_COL_ACTUAL_HOME, match.home_score, result_fill, result_font)
            put(row, consts.EXPORT_COL_ACTUAL_AWAY, match.away_score, result_fill, result_font)

        match_preds = predictions.get(match.id, {})
        match_scores = scores.get(match.id, {})
        for m in memberships:
            h, a, p = _member_columns(member_index[m.id])
            pred = match_preds.get(m.id)
            if pred:
                put(row, h, pred.predicted_home, pred_fill, pred_font)
                put(row, a, pred.predicted_away, pred_fill, pred_font)
            sc = match_scores.get(m.id)
            if sc:
                put(row, p, float(sc.points), points_fill, points_font)
        row += 1

    _finish_layout(ws, border, len(memberships), len(matches))
    return wb


def league_xlsx_bytes(league, now=None):
    """Serialize the league's results workbook to .xlsx bytes."""
    buf = BytesIO()
    build_league_workbook(league, now=now).save(buf)
    return buf.getvalue()
