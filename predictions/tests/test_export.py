"""Tests for the per-league results .xlsx export (key-gated download).

Two things matter most here and each gets a dedicated test:
  * the spreadsheet matches the shared league template (title + member columns,
    totals row, one row per match);
  * predictions for matches that haven't locked yet are NEVER written — an
    upcoming pick must stay invisible to anyone who downloads the file.
"""
from datetime import timedelta
from io import BytesIO
from unittest import mock

from django.core.cache import cache
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient, APITestCase

from predictions import consts, export
from predictions.models import Membership, Prediction
from predictions.throttles import ExportThrottle

from .factories import join, make_competition, make_league, make_match, make_team, make_user


def _member_cols(index):
    base = consts.EXPORT_FIRST_MEMBER_COL + index * consts.EXPORT_COLS_PER_MEMBER
    return base, base + 1, base + 2


def _name_to_col(ws):
    """Map each member's display name (row 1) to its first column."""
    out = {}
    col = consts.EXPORT_FIRST_MEMBER_COL
    while True:
        val = ws.cell(row=consts.EXPORT_TITLE_ROW, column=col).value
        if val is None:
            break
        out[val] = col
        col += consts.EXPORT_COLS_PER_MEMBER
    return out


def _row_by_home(ws, home_name):
    """Find the match row whose home-team cell equals home_name."""
    for r in range(consts.EXPORT_FIRST_MATCH_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=consts.EXPORT_COL_HOME).value == home_name:
            return r
    raise AssertionError(f"no match row for home team {home_name!r}")


class ExportBuilderTests(TestCase):
    def setUp(self):
        self.now = timezone.now()
        self.comp = make_competition()
        self.owner = make_user(display_name="آلیس")
        self.league = make_league(self.comp, owner=self.owner, name="لیگ اکسل")
        self.bob = make_user(display_name="باب")
        self.carol = make_user(display_name="کارول")
        self.m_owner = Membership.objects.get(league=self.league, user=self.owner)
        self.m_bob = join(self.league, user=self.bob)
        self.m_carol = join(self.league, user=self.carol)

    def _build(self):
        wb = export.build_league_workbook(self.league, now=self.now)
        return wb.active

    def test_title_and_member_header_and_totals(self):
        # A finished match so Alice has a non-zero total to show in row 2.
        home = make_team(self.comp, name="آلمان")
        away = make_team(self.comp, name="فرانسه")
        m = make_match(self.comp, home=home, away=away,
                       kickoff=self.now - timedelta(hours=2))
        Prediction.objects.create(membership=self.m_owner, match=m,
                                  predicted_home=2, predicted_away=1)
        m.home_score, m.away_score = 2, 1
        m.save()  # finished -> signal scores everyone (Alice exact = 10)

        ws = self._build()
        self.assertEqual(ws.cell(row=consts.EXPORT_TITLE_ROW,
                                 column=consts.EXPORT_COL_HOME).value, "لیگ اکسل")

        cols = _name_to_col(ws)
        self.assertEqual(set(cols), {"آلیس", "باب", "کارول"})
        # Owner joined first -> leftmost member column (stable join order).
        self.assertEqual(cols["آلیس"], consts.EXPORT_FIRST_MEMBER_COL)

        total_row = consts.EXPORT_TOTAL_ROW
        self.assertEqual(ws.cell(row=total_row, column=cols["آلیس"]).value, 10.0)
        self.assertEqual(ws.cell(row=total_row, column=cols["باب"]).value, 0.0)
        self.assertEqual(ws.cell(row=total_row, column=cols["کارول"]).value, 0.0)

    def test_finished_match_shows_result_predictions_and_points(self):
        home = make_team(self.comp, name="برزیل")
        away = make_team(self.comp, name="آرژانتین")
        m = make_match(self.comp, home=home, away=away,
                       kickoff=self.now - timedelta(hours=2))
        Prediction.objects.create(membership=self.m_owner, match=m,
                                  predicted_home=2, predicted_away=1)   # exact -> 10
        Prediction.objects.create(membership=self.m_bob, match=m,
                                  predicted_home=1, predicted_away=1)   # wrong -> 2
        # Carol does not predict.
        m.home_score, m.away_score = 2, 1
        m.save()

        ws = self._build()
        cols = _name_to_col(ws)
        r = _row_by_home(ws, "برزیل")

        self.assertEqual(ws.cell(row=r, column=consts.EXPORT_COL_AWAY).value, "آرژانتین")
        self.assertEqual(ws.cell(row=r, column=consts.EXPORT_COL_ACTUAL_HOME).value, 2)
        self.assertEqual(ws.cell(row=r, column=consts.EXPORT_COL_ACTUAL_AWAY).value, 1)

        ah, aa, ap = cols["آلیس"], cols["آلیس"] + 1, cols["آلیس"] + 2
        self.assertEqual(ws.cell(row=r, column=ah).value, 2)
        self.assertEqual(ws.cell(row=r, column=aa).value, 1)
        self.assertEqual(ws.cell(row=r, column=ap).value, 10.0)

        bh, ba, bp = cols["باب"], cols["باب"] + 1, cols["باب"] + 2
        self.assertEqual(ws.cell(row=r, column=bh).value, 1)
        self.assertEqual(ws.cell(row=r, column=ba).value, 1)
        self.assertEqual(ws.cell(row=r, column=bp).value, 2.0)

        # Carol predicted nothing: pick cells blank, but she still "played" at 0.
        ch, ca, cp = cols["کارول"], cols["کارول"] + 1, cols["کارول"] + 2
        self.assertIsNone(ws.cell(row=r, column=ch).value)
        self.assertIsNone(ws.cell(row=r, column=ca).value)
        self.assertEqual(ws.cell(row=r, column=cp).value, 0.0)

    def test_upcoming_predictions_are_invisible(self):
        # Match still open (kickoff well in the future) -> nobody's pick may show,
        # even though predictions exist in the database.
        home = make_team(self.comp, name="اسپانیا")
        away = make_team(self.comp, name="ایتالیا")
        m = make_match(self.comp, home=home, away=away,
                       kickoff=self.now + timedelta(hours=2))
        Prediction.objects.create(membership=self.m_owner, match=m,
                                  predicted_home=5, predicted_away=0)
        Prediction.objects.create(membership=self.m_bob, match=m,
                                  predicted_home=4, predicted_away=4)

        ws = self._build()
        cols = _name_to_col(ws)
        r = _row_by_home(ws, "اسپانیا")

        # Fixture is listed, but every member's three cells are empty.
        self.assertEqual(ws.cell(row=r, column=consts.EXPORT_COL_AWAY).value, "ایتالیا")
        self.assertIsNone(ws.cell(row=r, column=consts.EXPORT_COL_ACTUAL_HOME).value)
        for name in ("آلیس", "باب", "کارول"):
            for off in range(consts.EXPORT_COLS_PER_MEMBER):
                self.assertIsNone(ws.cell(row=r, column=cols[name] + off).value,
                                  f"{name} pick leaked for an upcoming match")

    def test_locked_match_reveals_picks_but_not_points(self):
        # Kicked off (so locked) but no result entered yet: picks show, points don't.
        home = make_team(self.comp, name="پرتغال")
        away = make_team(self.comp, name="هلند")
        m = make_match(self.comp, home=home, away=away,
                       kickoff=self.now - timedelta(minutes=10))
        Prediction.objects.create(membership=self.m_owner, match=m,
                                  predicted_home=3, predicted_away=0)

        ws = self._build()
        cols = _name_to_col(ws)
        r = _row_by_home(ws, "پرتغال")

        ah, aa, ap = cols["آلیس"], cols["آلیس"] + 1, cols["آلیس"] + 2
        self.assertEqual(ws.cell(row=r, column=ah).value, 3)
        self.assertEqual(ws.cell(row=r, column=aa).value, 0)
        self.assertIsNone(ws.cell(row=r, column=ap).value)          # not finished
        self.assertIsNone(ws.cell(row=r, column=consts.EXPORT_COL_ACTUAL_HOME).value)

    def test_hidden_predictions_show_result_and_total_but_no_picks(self):
        # Owner turned reveal_predictions off -> the result and the public total
        # still show, but every member's per-match pick/point cell stays blank
        # (same as the in-app match page).
        self.league.reveal_predictions = False
        self.league.save()
        home = make_team(self.comp, name="آلمان")
        away = make_team(self.comp, name="فرانسه")
        m = make_match(self.comp, home=home, away=away,
                       kickoff=self.now - timedelta(hours=2))
        Prediction.objects.create(membership=self.m_owner, match=m,
                                  predicted_home=2, predicted_away=1)  # exact -> 10
        m.home_score, m.away_score = 2, 1
        m.save()

        ws = self._build()
        cols = _name_to_col(ws)
        r = _row_by_home(ws, "آلمان")

        # Result is public...
        self.assertEqual(ws.cell(row=r, column=consts.EXPORT_COL_ACTUAL_HOME).value, 2)
        # ...and so is the running total (the leaderboard)...
        self.assertEqual(ws.cell(row=consts.EXPORT_TOTAL_ROW, column=cols["آلیس"]).value, 10.0)
        # ...but the per-match pick and points are hidden.
        for off in range(consts.EXPORT_COLS_PER_MEMBER):
            self.assertIsNone(ws.cell(row=r, column=cols["آلیس"] + off).value,
                              "reveal_predictions=False must hide the per-match pick/points")

    def test_empty_league_builds_without_error(self):
        # No members, no matches: the builder must still produce a valid sheet.
        Membership.objects.filter(league=self.league).delete()
        ws = export.build_league_workbook(self.league, now=self.now).active
        self.assertEqual(ws.cell(row=consts.EXPORT_TITLE_ROW,
                                 column=consts.EXPORT_COL_HOME).value, self.league.name)
        self.assertEqual(ws.freeze_panes, "E3")
        self.assertTrue(ws.sheet_view.rightToLeft)

    def test_result_entered_before_lock_stays_hidden(self):
        # Anomalous but possible: a result is entered while the match is still
        # before its lock time (kickoff in the future). Reveal is keyed off the
        # lock time, so the pick, the result AND the points must all stay hidden —
        # and the leaked points must not appear in the member's total either.
        home = make_team(self.comp, name="مراکش")
        away = make_team(self.comp, name="کرواسی")
        m = make_match(self.comp, home=home, away=away,
                       kickoff=self.now + timedelta(hours=2))   # lock is still future
        Prediction.objects.create(membership=self.m_owner, match=m,
                                  predicted_home=2, predicted_away=1)
        m.home_score, m.away_score = 2, 1
        m.save()  # finished early -> a MatchScore (10 pts) now exists for Alice

        ws = self._build()
        cols = _name_to_col(ws)
        r = _row_by_home(ws, "مراکش")

        # Row shows only the fixture.
        self.assertIsNone(ws.cell(row=r, column=consts.EXPORT_COL_ACTUAL_HOME).value)
        for off in range(consts.EXPORT_COLS_PER_MEMBER):
            self.assertIsNone(ws.cell(row=r, column=cols["آلیس"] + off).value,
                              "an early result must not reveal the pick/points pre-lock")
        # ...and the hidden points do not leak into the total.
        self.assertEqual(
            ws.cell(row=consts.EXPORT_TOTAL_ROW, column=cols["آلیس"]).value, 0.0)

    def test_cells_are_colour_coded(self):
        home = make_team(self.comp, name="نروژ")
        away = make_team(self.comp, name="بلژیک")
        m = make_match(self.comp, home=home, away=away,
                       kickoff=self.now - timedelta(hours=2))
        Prediction.objects.create(membership=self.m_owner, match=m,
                                  predicted_home=1, predicted_away=0)
        m.home_score, m.away_score = 1, 0
        m.save()

        ws = self._build()
        cols = _name_to_col(ws)
        r = _row_by_home(ws, "نروژ")

        def bg(cell):
            return cell.fill.fgColor.rgb  # e.g. "00800B02" — alpha-padded

        # Title banner, gold member header, amber team, navy result, grey points.
        self.assertTrue(bg(ws.cell(row=consts.EXPORT_TITLE_ROW,
                                   column=consts.EXPORT_COL_HOME)).endswith(consts.EXPORT_COLOR_TITLE_BG))
        self.assertTrue(bg(ws.cell(row=consts.EXPORT_TITLE_ROW,
                                   column=cols["آلیس"])).endswith(consts.EXPORT_COLOR_HEADER_BG))
        self.assertTrue(bg(ws.cell(row=r, column=consts.EXPORT_COL_HOME)).endswith(consts.EXPORT_COLOR_TEAM_BG))
        self.assertTrue(bg(ws.cell(row=r, column=consts.EXPORT_COL_ACTUAL_HOME)).endswith(consts.EXPORT_COLOR_RESULT_BG))
        self.assertTrue(bg(ws.cell(row=r, column=cols["آلیس"] + 2)).endswith(consts.EXPORT_COLOR_POINTS_BG))

        # Readability extras: frozen header/fixture pane + RTL view.
        self.assertEqual(ws.freeze_panes, "E3")
        self.assertTrue(ws.sheet_view.rightToLeft)

    def test_undecided_knockout_uses_bracket_label(self):
        m = make_match(self.comp, stage=consts.Stage.FINAL,
                       kickoff=self.now + timedelta(days=3))
        # No teams yet: bracket-slot labels carry the fixture.
        m.home_team = m.away_team = None
        m.home_label, m.away_label = "Match 101 Winner", "Match 102 Winner"
        m.save()

        ws = self._build()
        # The Persian bracket label is written, not a blank cell.
        labels = [ws.cell(row=r, column=consts.EXPORT_COL_HOME).value
                  for r in range(consts.EXPORT_FIRST_MATCH_ROW, ws.max_row + 1)]
        self.assertIn(consts.bracket_label_fa("Match 101 Winner"), labels)


class ExportEndpointTests(APITestCase):
    def setUp(self):
        self.comp = make_competition()
        self.owner = make_user(display_name="مدیر")
        self.league = make_league(self.comp, owner=self.owner, name="لیگ خروجی")

    def test_download_with_valid_key_is_public(self):
        # No authentication: the key alone is the credential.
        client = APIClient()
        url = reverse("api_export_league", args=[self.league.export_key])
        res = client.get(url)
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res["Content-Type"], consts.EXPORT_CONTENT_TYPE)
        self.assertIn("attachment", res["Content-Disposition"])

        wb = load_workbook(BytesIO(res.content))
        self.assertEqual(
            wb.active.cell(row=consts.EXPORT_TITLE_ROW,
                           column=consts.EXPORT_COL_HOME).value,
            "لیگ خروجی",
        )

    def test_invalid_key_returns_404(self):
        res = APIClient().get(reverse("api_export_league", args=["not-a-real-key"]))
        self.assertEqual(res.status_code, 404)

    def test_endpoint_hides_upcoming_predictions(self):
        bob = make_user(display_name="باب")
        m_bob = join(self.league, user=bob)
        home = make_team(self.comp, name="کرواسی")
        away = make_team(self.comp, name="بلژیک")
        m = make_match(self.comp, home=home, away=away,
                       kickoff=timezone.now() + timedelta(hours=3))
        Prediction.objects.create(membership=m_bob, match=m,
                                  predicted_home=2, predicted_away=2)

        res = APIClient().get(reverse("api_export_league", args=[self.league.export_key]))
        ws = load_workbook(BytesIO(res.content)).active
        r = _row_by_home(ws, "کرواسی")
        cols = _name_to_col(ws)
        for off in range(consts.EXPORT_COLS_PER_MEMBER):
            self.assertIsNone(ws.cell(row=r, column=cols["باب"] + off).value)

    def test_league_detail_exposes_export_key_and_url(self):
        client = APIClient()
        client.force_authenticate(user=self.owner)
        res = client.get(reverse("api_league_detail", args=[self.league.slug]))
        body = res.json()
        self.assertEqual(body["export_key"], self.league.export_key)
        self.assertTrue(
            body["export_url"].endswith(f"/api/export/{self.league.export_key}.xlsx")
        )

    def test_export_key_is_unique_per_league(self):
        other = make_league(self.comp, owner=make_user(), name="لیگ دیگر")
        self.assertNotEqual(self.league.export_key, other.export_key)
        self.assertTrue(self.league.export_key)

    def test_export_is_rate_limited(self):
        cache.clear()  # throttle history lives in the cache; isolate this test
        url = reverse("api_export_league", args=[self.league.export_key])
        client = APIClient()
        # DRF binds THROTTLE_RATES at import time, so patch the rate dict directly.
        with mock.patch.dict(ExportThrottle.THROTTLE_RATES,
                             {consts.THROTTLE_SCOPE_EXPORT: "1/min"}):
            first = client.get(url)
            second = client.get(url)
        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 429)
